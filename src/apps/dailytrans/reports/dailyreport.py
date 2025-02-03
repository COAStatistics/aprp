import calendar
import datetime
from collections import OrderedDict
from pathlib import Path
from typing import List, Optional, Dict, Union

import numpy as np
import openpyxl
import pandas as pd
from django.conf import settings
from django.db.models import QuerySet
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import create_engine
from sqlalchemy.engine import Engine

from apps.configs.models import Source, AbstractProduct
from apps.dailytrans.models import DailyTran, DailyTranQuerySet
from apps.dailytrans.utils import get_group_by_date_query_set
from apps.flowers.models import Flower
from apps.fruits.models import Fruit
from apps.watchlists.models import Watchlist, WatchlistItem, MonitorProfile


TEMPLATE = str(settings.BASE_DIR('apps/dailytrans/reports/template.xlsx'))

SHEET_FILL = PatternFill(
    start_color='f2dcdb',
    end_color='f2dcdb',
    fill_type='solid',
)

WEEKDAY = {
    0: '週一',
    1: '週二',
    2: '週三',
    3: '週四',
    4: '週五',
    5: '週六',
    6: '週日',
}

SHEET_FORMAT = {
    'F': '#,##0.0_);[Red]\\(#,##0.0\\)',
    'G': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'H': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'L': '0.0_ ',
    'M': '#,##0.0_ ',
    'N': '#,##0.0_ ',
    'O': '#,##0.0_ ',
    'P': '#,##0.0_ ',
    'Q': '#,##0.0_ ',
    'R': '#,##0.0_ ',
    'S': '#,##0.0_ ',
    'T': '#,##0_ ',
    'U': '#,##0.0_ ',
    'W': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'X': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'Y': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'Z': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'AA': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'AB': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'AC': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    'AD': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
}

# 日報下方資料來源對應品項及其來源說明;香蕉排在本項最後一列加上"一般農產品的資料來源說明",下載日報會出現欄位長度超過列印邊界,移到本項第一列以避免此問題
desc_1 = [
    ('香蕉',
     '香蕉產地價格(上品-中寮、中埔、旗山、美濃及高樹等農會查報上品價格之簡單平均)、(下品-中寮及中埔農會查報下品價格之簡單平均)'),
    ('落花生', '落花生產地價格(芳苑、虎尾、土庫、北港及元長等農會查報價格之簡單平均)'),
    ('紅豆', '紅豆產地價格(屏東縣萬丹鄉、新園鄉等產地農會查報價格之簡單平均)'),
    ('大蒜', '乾蒜頭產地價格(伸港、虎尾、土庫、元長及四湖等農會查報價格之簡單平均)'),
    ('甘薯', '甘藷產地價格(大城及水林等農會查報價格之簡單平均)'),
    ('桶柑', '桶柑產地價格(峨眉、寶山及和平等農會查報價格之簡單平均)'),
    ('柚子', '文旦柚產地價格(八里、冬山、西湖、斗六、麻豆、下營及瑞穗等農會查報價格之簡單平均)'),
    ('蓮霧', '蓮霧產地價格(六龜、枋寮及南州等農會查報價格之簡單平均)'),
    ('鳳梨', '金鑽鳳梨產地價格(名間、古坑、民雄、關廟、大樹、高樹和內埔等農會查報價格之簡單平均)'),
    ('雜柑', '檸檬產地價格(旗山、九如、里港、鹽埔及高樹等農會查報價格之簡單平均)'),
    ('梅', '竿採梅產地價格(國姓、六龜、甲仙和東河等農會查報價格之簡單平均)'),
    ('鳳梨釋迦', '鳳梨釋迦產地價格(台東地區農會查報價格)')
]

desc_2 = ['新興梨', '豐水梨', '柿子']

desc_3 = [
    '花卉交易(全部花卉市場) 交易量：火鶴花以單枝，文心蘭10枝，香水百合5枝─農產品行情報導，本部農糧署。',
    '95KG至155KG規格毛豬拍賣價格及土番鴨、白肉雞、雞蛋產地價格—中央畜產會。',
    '550KG以上肉牛產地價格—本部畜產試驗所恆春分所及中央畜產會。',
    '努比亞雜交閹公羊拍賣價格—彰化縣肉品拍賣市場。',
    '紅羽土雞產地價格(北區)—中華民國養雞協會。',
    '水產品產地價格─本部漁業署。'
]


def get_avg_price(qs, has_volume, has_weight):
    if has_volume and has_weight:  # 新增有日均重量的品項計算平均價格公式
        total_price = qs['avg_price'] * qs['sum_volume'] * qs['avg_avg_weight']
        total_volume_weight = qs['sum_volume'] * qs['avg_avg_weight']
        return total_price.sum() / total_volume_weight.sum() if len(total_volume_weight) else 0
    elif has_volume:
        total_price = qs['avg_price'] * qs['sum_volume']
        total_volume = qs['sum_volume']

        return total_price.sum() / total_volume.sum() if len(total_volume) else 0
    else:
        return qs['avg_price'].mean()


def get_avg_volume(qs):
    return qs['sum_volume'].mean() if not pd.isna(qs['sum_volume'].mean()) else 0


class LegacyDailyReportFactory(object):
    def __init__(self, specify_day):
        self.specify_day = specify_day
        self.this_week_start = self.specify_day - datetime.timedelta(6)
        self.this_week_end = self.specify_day
        self.last_week_start = self.this_week_start - datetime.timedelta(7)
        self.last_week_end = self.this_week_start - datetime.timedelta(1)
        self.last_year_month_start = datetime.datetime(self.specify_day.year - 1, self.specify_day.month, 1)
        self.last_year_month_end = datetime.datetime(self.specify_day.year - 1,
                                                     self.specify_day.month,
                                                     calendar.monthrange(self.specify_day.year - 1,
                                                                         self.specify_day.month)[1])

        self.row_visible = []
        self.row_marked = []
        self.result = {}
        self.col_dict = {}
        self.generate_list_dict()
        self.item_desc = []

    def generate_list_dict(self):
        for i in range(7):
            date = self.this_week_start.date() + datetime.timedelta(i)
            self.col_dict[f'{date}'] = f'{chr(77 + i)}'
            if i < 3:
                self.col_dict[f'{date}_volume'] = f'{chr(88 + i)}'
            else:
                self.col_dict[f'{date}_volume'] = f'A{chr(65 + i - 3)}'

    def check_months(self, item):
        # 不在監控品項月份變更底色改為在顯示月份內的品項顯示        
        # if item.months.filter(name__icontains=self.specify_day.month) or item.always_display:
        # 原判斷條件導致10/11/12月份的品項在1/2月分也會出現
        if item.months.filter(name=f'{self.specify_day.month}月') or item.always_display:
            self.row_visible.append(item.row)
            if item.product.name == '梨':
                if self.specify_day.month in [5, 6]:
                    self.item_desc.append('豐水梨')
                elif self.specify_day.month in [7, 8]:
                    self.item_desc.append('新興梨')
            else:
                self.item_desc.append(item.product.name)

    def input_sheet_date(self, sheet, index):
        month = (self.this_week_start + datetime.timedelta(index)).month
        day = (self.this_week_start + datetime.timedelta(index)).day
        sym = chr(77 + index)
        sheet[f'{sym}8'] = f'{month}月\n{day}日'
        sym = f'A{chr(65 + index - 3)}' if index >= 3 else chr(88 + index)
        sheet[f'{sym}8'] = f'{month}月\n{day}日'
        return sheet

    def get_data(self, query_set, product, row, monitor_price):
        qs, has_volume, has_weight = get_group_by_date_query_set(query_set, self.last_week_start, self.this_week_end)
        self.result[product] = {}
        for i, q in qs.iterrows():

            if q['date'] >= self.this_week_start.date():
                self.result[product].update(
                    {f"""{self.col_dict[f"{q['date']}"]}{row}""": q['avg_price']}
                )
            if has_volume and q['date'] >= self.this_week_start.date():
                self.result[product].update(
                    {f"""{self.col_dict[f"{q['date']}_volume"]}{row}""": q[
                        'sum_volume'
                    ]
                     })

        last_qs = qs[(pd.to_datetime(qs['date']).dt.date >= self.last_week_start.date())
                     & (pd.to_datetime(qs['date']).dt.date <= self.last_week_end.date())]
        this_qs = qs[(pd.to_datetime(qs['date']).dt.date >= self.this_week_start.date())
                     & (pd.to_datetime(qs['date']).dt.date <= self.this_week_end.date())]

        last_avg_price = get_avg_price(last_qs, has_volume, has_weight)
        this_avg_price = get_avg_price(this_qs, has_volume, has_weight)
        if last_avg_price > 0:
            self.result[product].update(
                {f'L{row}': (this_avg_price - last_avg_price) / last_avg_price * 100}
            )
        if has_volume:
            last_avg_volume = get_avg_volume(last_qs)
            this_avg_volume = get_avg_volume(this_qs)
            self.result[product].update({f'T{row}': this_avg_volume})
            if last_avg_volume > 0:
                self.result[product].update(
                    {f'U{row}': (this_avg_volume - last_avg_volume) / last_avg_volume * 100}
                )
        if monitor_price:
            self.result[product].update({f'F{row}': monitor_price})
        self.result[product].update(
            {f'H{row}': this_avg_price, f'W{row}': last_avg_price}
        )

    def update_data(self, query_set, product, row):
        qs, has_volume, has_weight = get_group_by_date_query_set(query_set)
        last_year_avg_price = get_avg_price(qs, has_volume, has_weight)
        if last_year_avg_price > 0:
            self.result[product].update({f'G{row}': last_year_avg_price})

    def update_rams(self, item, row):
        query_set = DailyTran.objects.filter(product__in=item.product_list(), source__in=item.sources())
        for i in range(7):
            week_day = self.this_week_start + datetime.timedelta(i)
            qs = query_set.filter(date__lte=week_day.date()).order_by('-date')
            if qs:
                qs = qs.first()
            else:
                continue
            if i == 0 or qs.date == week_day.date():
                self.result[item.product.name].update(
                    {f"{self.col_dict[f'{week_day.date()}']}{row}": qs.avg_price}
                )
            else:
                self.result[item.product.name].update(
                    {
                        f"{self.col_dict[f'{week_day.date()}']}{row}": qs.date.strftime(
                            '(%m/%d)'
                        )
                    }
                )

    def update_cattles(self, item, row):
        query_set = DailyTran.objects.filter(product__in=item.product_list())
        last_week_price = []
        this_week_price = []
        for i in range(7):
            week_day = self.this_week_start + datetime.timedelta(i)
            qs = query_set.filter(date__lte=week_day.date()).order_by('-date')
            if not qs:
                continue
            qs = qs.first()
            this_week_price.append(qs.avg_price)
            self.result[item.product.name].update(
                {f"{self.col_dict[f'{week_day.date()}']}{row}": qs.avg_price}
            )
        for i in range(1, 8):
            week_day = self.this_week_start - datetime.timedelta(i)
            qs = query_set.filter(date__lte=week_day.date()).order_by('-date')
            if qs:
                qs = qs.first()
                last_week_price.append(qs.avg_price)
        if len(last_week_price):
            last_week_avg_price = sum(last_week_price) / len(last_week_price)
            self.result[item.product.name].update({f'W{row}': last_week_avg_price})
        if len(this_week_price):
            this_week_avg_price = sum(this_week_price) / len(this_week_price)
            self.result[item.product.name].update({f'H{row}': this_week_avg_price})
        if len(last_week_price) and len(this_week_price):
            self.result[item.product.name].update(
                {
                    f'L{row}': (this_week_avg_price - last_week_avg_price)
                               / last_week_avg_price
                               * 100
                }
            )

    def report(self):
        watchlist = Watchlist.objects.filter(
            start_date__year=self.specify_day.year,
            start_date__month__lte=self.specify_day.month,
            end_date__month__gte=self.specify_day.month
        ).first()
        monitor = MonitorProfile.objects.filter(watchlist=watchlist, row__isnull=False)

        for item in monitor:
            # self.row_visible.append(item.row)
            query_set = DailyTran.objects.filter(product__in=item.product_list())

            # 因應措施是梨
            if self.specify_day.month in [5, 6]:
                if item.product.id == 50182:
                    # 56只抓豐水梨 50186
                    query_set = query_set.filter(product=50186)
            elif self.specify_day.month in [7, 8]:
                if item.product.id == 50182:
                    # 78只抓新興梨 50185
                    query_set = query_set.filter(product=50185)

            if item.sources():
                query_set = query_set.filter(source__in=item.sources())
            self.get_data(query_set, item.product.name, item.row, item.price)
            # last year avg_price of month
            query_set = DailyTran.objects.filter(product__in=item.product_list(),
                                                 date__year=self.specify_day.year - 1,
                                                 date__month=self.specify_day.month)

            # 因應措施是梨
            if self.specify_day.month in [5, 6]:
                if item.product.id == 50182:
                    # 56只抓豐水梨 50186
                    query_set = query_set.filter(product=50186)
            elif self.specify_day.month in [7, 8]:
                if item.product.id == 50182:
                    # 78只抓新興梨 50185
                    query_set = query_set.filter(product=50185)
            if item.sources():
                query_set = query_set.filter(source__in=item.sources())
            self.update_data(query_set, item.product.name, item.row)

            if '羊' in item.product.name:
                self.update_rams(item, item.row)
            if '牛' in item.product.name:
                self.update_cattles(item, item.row)
            self.check_months(item)

        # 長糯, 稻穀, 全部花卉 L, 火鶴花 FB, 文心蘭 FO3
        extra_product = [(3001, 10), (3002, 9), (3508, 99), (3509, 100), (3510, 103)]
        for item in extra_product:
            self._extract_data(item[1], WatchlistItem, item[0], None, self.specify_day)

        # 香蕉台北一二批發
        self._extract_data(73, Fruit, 50063, Source.objects.filter(id__in=[20001, 20002]), self.specify_day)

        # 青香蕉下品()內銷)
        self._extract_data(72, Fruit, 59019, Source.objects.filter(id__in=range(10030, 20001)), self.specify_day)

        # 2020/4/16 主管會報陳副主委要求花卉品項,農糧署建議新增香水百合 FS  
        self._extract_data(107, Flower, 60068, Source.objects.filter(id__in=[30001, 30002, 30003, 30004, 30005]),
                           self.specify_day)

    def _extract_data(self, row, model, product_id, sources=None, date=None):
        self.row_visible.append(row)
        if model == WatchlistItem:
            watchlist_item = model.objects.filter(id=product_id).first()
            product = watchlist_item.product
            if watchlist_item.sources.all():
                sources = watchlist_item.sources.all()
        else:
            product = model.objects.get(id=product_id)
        query_set = DailyTran.objects.filter(product=product)
        if sources:
            query_set = query_set.filter(source__in=sources)
        self.get_data(query_set, f'{product.name}{product.type}', row, None)
        if date:
            query_set = DailyTran.objects.filter(product=product, date__year=date.year - 1, date__month=date.month)
            if sources:
                query_set = query_set.filter(source__in=sources)
            self.update_data(query_set, f'{product.name}{product.type}', row)

    @staticmethod
    def get_sheet_format(key):
        chr1 = key[0]
        chr2 = key[1]
        if 65 <= ord(chr1) <= 90 and 65 <= ord(chr2) <= 90:
            return SHEET_FORMAT.get(chr1 + chr2)
        else:
            return SHEET_FORMAT.get(chr1)

    @staticmethod
    def roc_date_format(date, sep='.'):
        return sep.join([
            str(date.year - 1911),
            str(date.month).zfill(2),
            str(date.day).zfill(2)
        ])

    def get_sheet(self):
        wb = openpyxl.load_workbook(TEMPLATE)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]

        for i in range(7):
            sheet = self.input_sheet_date(sheet, i)
        sheet['G6'] = (
            f'{self.specify_day.year - 1912}\n年{self.specify_day.month}月\n平均價格'
        )
        last_week_range = f'{self.last_week_start.month}/{self.last_week_start.day}~{self.last_week_end.month}/{self.last_week_end.day}'
        sheet['W8'] = last_week_range

        for key, value in self.result.items():
            for k, v in value.items():
                try:
                    if v != 0 or v == 0 and 'L' in k or v == 0 and 'U' in k:
                        sheet[k] = v
                    sheet_format = self.get_sheet_format(k)
                    if sheet_format:
                        sheet[k].number_format = sheet_format
                except Exception:
                    pass

        for i in range(9, 132):
            if i not in self.row_visible:
                sheet.row_dimensions[i].hidden = True
            # 依袁麗惠要求,日報取消品項底色識別
            # 第二階段隱藏品項欄位
            # td = sheet.cell(row=i, column=2)
            # A.品項欄位去除底色
            # td.fill = PatternFill(
            #     fill_type='solid',
            #     start_color='FFFFFF',
            #     end_color='FFFFFF'
            # )

        # 第二階段隱藏品項欄位後日報下方說明欄,依品項顯示月份對應調整資料來源文字說明處理
        for rows in sheet['A133:U148']:
            for cell in rows:
                # 資料來源字型統一為標楷體
                cell.font = Font(name='標楷體', size=13)
                row_no = cell.row
                if row_no > 134:
                    cell.value = None

        sheet.cell(row=133, column=1).value = sheet.cell(row=133, column=1).value.replace('本會', '本部')
        sheet.cell(row=134, column=1).value = sheet.cell(row=134, column=1).value.replace('本會', '本部')

        now_row = 135
        # 一般農產品的資料來源說明欄位處理
        for i in desc_1:
            item_name = i[0]
            desc_1_text = i[1]

            # append_desc
            if item_name in self.item_desc:
                td = sheet.cell(row=now_row, column=1)
                tmp = (now_row == 135 and '3.') or '   '
                td.value = f"{tmp}{desc_1_text}；"
                now_row += 1
        td = sheet.cell(row=now_row - 1, column=1)
        td.value = td.value.replace('；', '—農產品價格查報，本部農糧署。')
        desc_2_tmp = []
        pn = 4
        for item_name in desc_2:
            if item_name in self.item_desc:
                if item_name == '柿子':
                    desc_2_tmp.append('甜柿')
                else:
                    desc_2_tmp.append(item_name)
        if desc_2_tmp:
            td = sheet.cell(row=now_row, column=1)
            desc_2_text = '4.' + '、'.join(desc_2_tmp) + \
                          '交易量價(東勢果菜市場價格)－農產品行情報導，本部農糧署。'
            td.value = desc_2_text
            now_row += 1
            pn += 1
        # 其餘花卉,畜禽,水產類的資料來源說明欄位處理
        for p in desc_3:
            td = sheet.cell(row=now_row, column=1)
            td.value = f"{str(pn)}.{p}"
            pn += 1
            now_row += 1

        return wb

    def __call__(self, output_dir=settings.BASE_DIR):
        date = self.specify_day + datetime.timedelta(1)
        file_name = f'{self.roc_date_format(self.this_week_start)}-{self.roc_date_format(self.this_week_end)}價格{WEEKDAY.get(date.weekday())}.xlsx'

        file_path = Path(output_dir, file_name)

        self.report()
        sheet = self.get_sheet()
        sheet.save(file_path)

        return file_name, file_path


class Database:
    __db = None

    def __init__(self):
        # 連結資料庫相關設定
        database = settings.DATABASES['default']['NAME']
        user = settings.DATABASES['default']['USER']
        password = settings.DATABASES['default']['PASSWORD']
        host = settings.DATABASES['default']['HOST']
        port = settings.DATABASES['default']['PORT']
        app_name = 'pd.read_sql'

        self.__conn = create_engine(
            f'postgresql://{user}:{password}@{host}:{port}/{database}?application_name={app_name}',
            echo=False
        )

    @classmethod
    def get_db_connection(cls):
        if cls.__db is None:
            cls.__db = Database()
        else:
            try:
                cls.__db.__conn.execute("SELECT 1")
            except Exception as e:
                print("Database connection error: ", e)
                cls.__db = Database()

        return cls.__db.__conn


class QueryString:
    def __init__(self):
        self.base_query = """
                SELECT "dailytrans_dailytran"."source_id",
                       "dailytrans_dailytran"."avg_price",
                       "dailytrans_dailytran"."avg_weight",
                       "dailytrans_dailytran"."volume",
                       "dailytrans_dailytran"."date"
                FROM "dailytrans_dailytran"
                """
        self.__freeze_query = self.base_query
        self.__temp_query: Optional[str] = None
        self.dict_query = {}
        self.__dict_query = {}

    def add_where(self):
        self.base_query += ' WHERE ('

        return self

    @staticmethod
    def add_and_keyword(query_str:str, has_and: bool):
        return f' AND {query_str}' if has_and else query_str

    def add_where_by_product_in(self, products: List[AbstractProduct], has_and=True):
        key = 'product_list'
        q = self.add_and_keyword(f'"dailytrans_dailytran"."product_id" IN %({key})s ', has_and)
        self.base_query += q
        self.dict_query[key] = tuple(p.id for p in products)

        return self

    def add_where_by_product_id(self, product_id: int, has_and=True):
        key = 'product_id'
        q = self.add_and_keyword(f'"dailytrans_dailytran"."product_id" = %({key})s ', has_and)
        self.base_query += q
        self.dict_query[key] = product_id

        return self

    def add_where_by_source_in(self, monitor: MonitorProfile, has_and=True):
        sources = monitor.sources() if callable(monitor.sources) else monitor.sources

        if not sources:
            return

        key = 'source_list'
        q = self.add_and_keyword(f'"dailytrans_dailytran"."source_id" IN %({key})s ', has_and)
        self.base_query += q
        self.dict_query[key] = tuple(s.id for s in sources)

        return self

    def add_where_by_date_between(self, start_date: datetime.date, end_date: datetime.date, has_and=True):
        key_start_dt = 'start_date'
        key_end_dt = 'end_date'
        q = self.add_and_keyword(
            f'("dailytrans_dailytran"."date" between %({key_start_dt})s AND %({key_end_dt})s)',
            has_and
        )
        self.base_query += q
        self.dict_query[key_start_dt] = start_date
        self.dict_query[key_end_dt] = end_date

        return self

    def add_where_by_last_year_and_month(self, dt: datetime.date, has_and=True):
        last_year = dt.year - 1
        days = calendar.monthrange(last_year, dt.month)[1]
        start_date = datetime.datetime.strptime(f'{last_year}-{dt.month}-01', '%Y-%m-%d')
        end_date = datetime.datetime.strptime(f'{last_year}-{dt.month}-{days}', '%Y-%m-%d')

        self.add_where_by_date_between(start_date, end_date, has_and)

        return self

    def build(self):
        query = f'{self.base_query});'
        self.base_query = self.__freeze_query

        return query

    def save(self):
        self.__temp_query = self.base_query
        self.__dict_query = self.dict_query.copy()

        return self

    def load(self, keep=False):
        self.base_query = self.__temp_query
        self.dict_query = self.__dict_query.copy()

        if not keep:
            self.clean()

        return self

    def clean(self):
        self.__temp_query = None
        self.__dict_query.clear()

        return self


class DailyTranHandler:
    """
    此類別用於對日交易(DailyTran)資料做相關處理，為提升效能，全程使用 pandas 套件處理資料，並且
    採用直連資料庫的方式，而非透過 Django ORM(經過測試，效能上有明顯差異)。
    """

    __conn: Optional[Engine] = None

    def __init__(self, sql: str, params: dict):
        if self.__conn is None:
            self.__conn = Database.get_db_connection()

        self.df = pd.read_sql_query(sql=sql, con=self.__conn, params=params)

    @property
    def group_by_columns(self) -> List[str]:
        return ['date', 'avg_price', 'num_of_source', 'sum_volume', 'avg_avg_weight']

    @property
    def has_volume(self) -> bool:
        """
        確認資料是否有交易量(volume)
        """

        return (not self.df.empty) and (self.df['volume'].notna().sum() / self.df['avg_price'].count() > 0.8)

    @property
    def has_weight(self) -> bool:
        """
        確認資料是否有平均重量(avg_weight)
        """

        return (not self.df.empty) and (self.df['avg_weight'].notna().sum() / self.df['avg_price'].count() > 0.8)

    @property
    def fulfilled_df(self) -> pd.DataFrame:
        """
        針對日交易原始資料進行一些填充和計算，以便後續的資料處理
        """

        df = self.df.query('volume > 0 and avg_weight > 0') if self.has_volume and self.has_weight else self.df

        # 數據處理和計算:
        # 1. 將缺失值填充為 1，不用因為缺失值設定判斷式再進行計算
        # 2. 計算單日單一交易地點的總價格和總交易重量
        df = (
            df
            .assign(vol_for_calculation=lambda x: x['volume'].fillna(1))
            .assign(wt_for_calculation=lambda x: x['avg_weight'].fillna(1))
            .assign(avg_price=lambda x: x['avg_price'] * x['wt_for_calculation'] * x['vol_for_calculation'])
            .assign(wt_for_calculation=lambda x: x['wt_for_calculation'] * x['vol_for_calculation'])
        )

        # 若所有資料無來源 ID，則將來源 ID 填充為 1
        return df.assign(source_id=lambda x: x['source_id'].fillna(1)) if all(pd.isna(df['source_id'])) else df

    @property
    def df_with_group_by_date(self) -> pd.DataFrame:
        """
        針對日交易進行日期分組，並計算平均價格和平均重量
        """

        if self.df.empty:
            return pd.DataFrame(columns=self.group_by_columns)

        # 處理流程:
        # 1. 按日期和來源市場 ID 分組
        # 2. 再按日期分組
        # 3. 計算最終的平均價格(avg_price) 和 平均重量(avg_weight)
        # 4. 處理缺失的 總交易量(sum_volume) 和 每日平均重量(avg_avg_weight)
        df = (
            self
            .fulfilled_df
            .groupby(['date', 'source_id'])
            .sum()
            .assign(num_of_source=1)
            .groupby('date')
            .sum()
            .assign(avg_price=lambda x: x['avg_price'] / x['wt_for_calculation'])
            .assign(avg_weight=lambda x: x['wt_for_calculation'] / x['vol_for_calculation'])
            .reset_index()
            .sort_values('date')
            .rename(columns={'volume': 'sum_volume', 'avg_weight': 'avg_avg_weight'})
            .assign(sum_volume=lambda x: np.where(self.has_volume, x['vol_for_calculation'], x['num_of_source']))
            .assign(avg_avg_weight=lambda x: np.where(self.has_weight, x['avg_avg_weight'], 1))
        )

        return df[self.group_by_columns]
    
    def _get_df_query_by_date(
            self, start_date: Optional[datetime.date] = None, end_date: Optional[datetime.date] = None
    ) -> pd.DataFrame:
        """
        根據日期區間篩選日交易資料
        """

        return (
            self.df_with_group_by_date.query(
                f"@pd.to_datetime(date) >= '{start_date}' and "
                f"@pd.to_datetime(date) <= '{end_date}'"
            )
            if start_date and end_date
            else self.df_with_group_by_date
        )
    
    def get_avg_price(
            self, start_date: Optional[datetime.date] = None, end_date: Optional[datetime.date] = None
    ) -> Union[int, float]:
        """
        計算指定日期區間的平均價格
        """

        df = self._get_df_query_by_date(start_date, end_date)

        # 新增有日均重量的品項計算平均價格公式
        if self.has_volume and self.has_weight:
            total_price = df['avg_price'] * df['sum_volume'] * df['avg_avg_weight']
            total_volume_weight = df['sum_volume'] * df['avg_avg_weight']
            return total_price.sum() / total_volume_weight.sum() if len(total_volume_weight) else 0

        elif self.has_volume:
            total_price = df['avg_price'] * df['sum_volume']
            total_volume = df['sum_volume']

            return total_price.sum() / total_volume.sum() if len(total_volume) else 0
        else:
            return df['avg_price'].mean()

    def get_avg_volume(
            self, start_date: Optional[datetime.date] = None, end_date: Optional[datetime.date] = None
    ) -> Union[int, float]:
        """
        計算指定日期區間的平均交易量
        """

        df = self._get_df_query_by_date(start_date, end_date)

        return 0 if pd.isna(df['sum_volume'].mean()) else df['sum_volume'].mean()


class ExtraItem:
    EXTRA_ITEMS = [
        {
            'product_id': 15,
            'row': 10,
        },
        {
            'product_id': 19,
            'row': 9,
        },
        {
            'product_id': 30002,
            'row': 99,
        },
        {
            'product_id': 60051,
            'row': 100,
        },
        {
            'product_id': 60066,
            'row': 103,
        },
        {
            'product_id': 50063,
            'row': 73,
            'sources_id': [20001, 20002]
        },
        {
            'product_id': 59019,
            'row': 72,
            'sources_id': [i for i in range(10030, 10135)]
        },
        {
            'product_id': 60068,
            'row': 107,
            'sources_id': [30001, 30002, 30003, 30004, 30005]
        },
    ]
    
    def __init__(self, product_id: int, row: int, sources_id: Optional[List[int]] = None):
        self.product_id = product_id
        self.row = row
        self.sources_id = sources_id

    @classmethod
    def get_extra_monitors(cls) -> List[MonitorProfile]:
        _list = []

        for d in cls.EXTRA_ITEMS:
            item = ExtraItem(**d)
            product = AbstractProduct.objects.get(id=item.product_id)
            product.name = f'{product.name}{product.type}'
            watchlist = Watchlist.objects.last()
            monitor = MonitorProfile(product=product, watchlist=watchlist, type=product.type, row=item.row)

            if item.sources_id:
                monitor.sources = list(Source.objects.filter(id__in=item.sources_id))

            _list.append(monitor)

        return _list


class ExcelHandler:
    __DICT_WEEKDAY = {
        0: '週一',
        1: '週二',
        2: '週三',
        3: '週四',
        4: '週五',
        5: '週六',
        6: '週日',
    }

    __CELL_FORMAT = {
        'F': '#,##0.0_);[Red]\\(#,##0.0\\)',
        'G': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'H': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'L': '0.0_ ',
        'M': '#,##0.0_ ',
        'N': '#,##0.0_ ',
        'O': '#,##0.0_ ',
        'P': '#,##0.0_ ',
        'Q': '#,##0.0_ ',
        'R': '#,##0.0_ ',
        'S': '#,##0.0_ ',
        'T': '#,##0_ ',
        'U': '#,##0.0_ ',
        'W': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'X': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'Y': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'Z': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'AA': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'AB': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'AC': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
        'AD': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"?_-;_-@_-',
    }

    __DICT_CROP_DESC = OrderedDict({
        '落花生': '  落花生產地價格(芳苑、虎尾、土庫、北港及元長等農會查報價格之簡單平均)',
        '紅豆': '  紅豆產地價格(屏東縣萬丹鄉、新園鄉等產地農會查報價格之簡單平均)',
        '大蒜': '  乾蒜頭產地價格(伸港、虎尾、土庫、元長及四湖等農會查報價格之簡單平均)',
        '甘薯': '  甘藷產地價格(大城及水林等農會查報價格之簡單平均)',
        '桶柑': '  桶柑產地價格(峨眉、寶山及和平等農會查報價格之簡單平均)',
        '柚子': '  文旦柚產地價格(八里、冬山、西湖、斗六、麻豆、下營及瑞穗等農會查報價格之簡單平均)',

        # 114 年上半年移除蓮霧
        # '蓮霧': '  蓮霧產地價格(六龜、枋寮及南州等農會查報價格之簡單平均)',
        '鳳梨': '  金鑽鳳梨產地價格(名間、古坑、民雄、關廟、大樹、高樹和內埔等農會查報價格之簡單平均)',
        '雜柑': '  檸檬產地價格(旗山、九如、里港、鹽埔及高樹等農會查報價格之簡單平均)',
        '梅': '  竿採梅產地價格(國姓、六龜、甲仙和東河等農會查報價格之簡單平均)',
        '鳳梨釋迦': '  鳳梨釋迦產地價格(台東地區農會查報價格)',
    })

    __LIST_OTHER_DESC = [
        '花卉交易(全部花卉市場) 交易量：火鶴花以單枝，文心蘭10枝，香水百合5枝─農產品行情報導，本部農糧署。',
        '95KG至155KG規格毛豬拍賣價格及土番鴨、白肉雞、雞蛋產地價格—中央畜產會。',
        '550KG以上肉牛產地價格—本部畜產試驗所恆春分所及中央畜產會。',
        '努比亞雜交閹公羊拍賣價格—彰化縣肉品拍賣市場。',
        '紅羽土雞產地價格(北區)—中華民國養雞協會。',
        '水產品產地價格─本部漁業署。',
    ]

    TEMPLATE_PATH = str(settings.BASE_DIR('apps/dailytrans/reports/template.xlsx'))

    DESC_LIST_NUM = 4
    DESC_ROW_START = 137
    END_OF_CROP_DESC = '—農產品價格查報，本部農糧署。'
    SPECIFIED_CROP_DESC = '交易量價(東勢果菜市場價格)－農產品行情報導，本部農糧署。'

    ROW_NUM_START = 9
    ROW_NUM_END = 132

    def __init__(self, factory: 'DailyReportFactory', output_dir=settings.BASE_DIR):
        self.factory = factory
        self.output_dir = output_dir
        self.specify_day = factory.specify_day

        self.__visible_rows: List[int] = []
        self.__desc_product_names: List[str] = []
        self.__wb: Optional[openpyxl.Workbook] = None
        self.__sheet: Optional[Worksheet] = None
        self.__desc_row_num: Optional[int] = None
        self.__desc_list_num: Optional[int] = None

    @staticmethod
    def convert_to_roc_date_str(date, sep='.'):
        return sep.join([
            str(date.year - 1911),
            str(date.month).zfill(2),
            str(date.day).zfill(2)
        ])

    @property
    def dict_crop_desc(self):
        return self.__DICT_CROP_DESC

    @property
    def visible_rows(self):
        return self.__visible_rows

    @visible_rows.setter
    def visible_rows(self, row: int):
        self.__visible_rows.append(row)

    @property
    def desc_product_names(self):
        return self.__desc_product_names

    @desc_product_names.setter
    def desc_product_names(self, product_name: str):
        self.__desc_product_names.append(product_name)

    @property
    def wb(self):
        if self.__wb is None:
            self.__wb = openpyxl.load_workbook(self.TEMPLATE_PATH)
            self.__sheet = self.__wb[self.__wb.sheetnames[0]]
            self.__init_sheet()

        return self.__wb

    @property
    def sheet(self):
        return self.__sheet

    @property
    def desc_row_num(self):
        if self.__desc_row_num is None:
            self.__desc_row_num = self.DESC_ROW_START
        else:
            self.__desc_row_num += 1

        return self.__desc_row_num

    @property
    def desc_list_num(self):
        if self.__desc_list_num is None:
            self.__desc_list_num = self.DESC_LIST_NUM
        else:
            self.__desc_list_num += 1

        return self.__desc_list_num

    @property
    def file_name(self):
        today = self.specify_day + datetime.timedelta(1)
        last_week_date = self.convert_to_roc_date_str(self.factory.this_week_start)
        this_week_date = self.convert_to_roc_date_str(self.factory.this_week_end)

        return f'{last_week_date}-{this_week_date}價格{self.__DICT_WEEKDAY.get(today.weekday())}.xlsx'

    @property
    def file_path(self):
        return Path(self.output_dir, self.file_name)

    def __set_text_of_date(self):
        for i in range(7):
            month = (self.factory.this_week_start + datetime.timedelta(i)).month
            day = (self.factory.this_week_start + datetime.timedelta(i)).day

            # the chr(77) is 'M'
            sym = chr(77 + i)
            self.__sheet[f'{sym}8'] = f'{month}月\n{day}日'

            sym = f'A{chr(65 + i - 3)}' if i >= 3 else chr(88 + i)
            self.__sheet[f'{sym}8'] = f'{month}月\n{day}日'

    def __set_text_of_date_range_and_avg_price(self):
        roc_year = self.specify_day.year - 1911
        last_year = roc_year - 1
        month_of_last_year = self.specify_day.month
        self.__sheet['G6'] = f'{last_year}\n年{month_of_last_year}月\n平均價格'

        dt_start = self.factory.last_week_start
        dt_end = self.factory.this_week_end
        last_week_range_text = f'{dt_start.month}/{dt_start.day}~{dt_end.month}/{dt_end.day}'
        self.__sheet['W8'] = last_week_range_text

    def __set_product_values(self):
        if self.factory.result:
            for product_name, dict_values in self.factory.result.items():
                for cell_name, avg_price in dict_values.items():
                    # 'L' -> 與前一週比較(%(平均價格)), 'U' -> 與前一週比較(%(交易量))
                    if avg_price != 0 or 'L' in cell_name or 'U' in cell_name:
                        self.__sheet[cell_name] = avg_price

                    if self.get_cell_format(cell_name):
                        self.__sheet[cell_name].number_format = self.get_cell_format(cell_name)

    def __set_row_visible(self):
        for i in range(self.ROW_NUM_START, self.ROW_NUM_END):
            if i not in self.visible_rows:
                self.__sheet.row_dimensions[i].hidden = True

    def __set_crop_desc(self):
        for i in self.dict_crop_desc.values():
            self.__sheet.cell(row=self.desc_row_num, column=1).value = i

    def __set_specified_crop_desc(self):
        product_names = []

        for i in self.visible_rows:
            if i == 57:
                product_names.append('新興梨')
            elif i == 58:
                product_names.append('豐水梨')
            elif i == 67:
                product_names.append('甜柿')

        if product_names:
            text = f'{self.desc_list_num}.{"、".join(product_names)}{self.SPECIFIED_CROP_DESC}'
            self.__sheet.cell(row=self.desc_row_num, column=1).value = text

    def __set_other_desc(self):
        for i in self.__LIST_OTHER_DESC:
            self.__sheet.cell(row=self.desc_row_num, column=1).value = f'{self.desc_list_num}.{i}'

    def __init_sheet(self):
        self.__set_text_of_date()
        self.__set_text_of_date_range_and_avg_price()
        self.__set_product_values()
        self.__set_row_visible()
        self.__set_crop_desc()
        self.__set_specified_crop_desc()
        self.__set_other_desc()

    def get_cell_format(self, cell_name: str):
        alph1 = cell_name[0]
        alph2 = cell_name[1]

        return self.__CELL_FORMAT.get(f'{alph1}{alph2}') if alph2.isalpha() else self.__CELL_FORMAT.get(alph1)

    def remove_crop_desc(self, key: str):
        self.__DICT_CROP_DESC.pop(key)

    def save(self):
        self.wb.save(self.file_path)


class DailyReportFactory:
    """
    為了較好的可讀性與執行效能，將原先的日報表類別(DailyReportFactory)重構。
    """

    def __init__(self, specify_day: datetime.datetime):
        self.specify_day = specify_day

        """
        代表整個報表的結果資料，格式為:
        {
            '青香蕉(內銷)': {
              'M71': 60.8,
              'N71': 62.0,
              'O71': 62.4,
              'P71': 62.8,
              'Q71': 63.4,
              'S71': 63.6
              },
              ...
        }
        """
        self.result: Dict[str, Dict[str, float]] = {}
        self.query = QueryString()
        self.excel_handler = ExcelHandler(self)

        self.__watchlist: Optional[Watchlist] = None
        self.__monitor_profile_qs: Optional[DailyTranQuerySet] = None
        self.__col_mapping: Optional[Dict[str, str]] = None
        self.__monitor: Optional[MonitorProfile] = None
        self.__two_weeks_handler: Optional[DailyTranHandler] = None
        self.__df_two_weeks: Optional[pd.DataFrame] = None
        self.__visible_rows: List[int] = []

    def __call__(self):
        self.report()
        self.excel_handler.save()

        return self.excel_handler.file_name, self.excel_handler.file_path

    @property
    def col_mapping(self):
        """
        生成對應日期的 Excel 欄位(column)，以方便後續對應列(row)

        The dict will like this:
        {'2024-12-11': 'M',
         '2024-12-11_volume': 'X',
         '2024-12-12': 'N',
         '2024-12-12_volume': 'Y',
         '2024-12-13': 'O',
         '2024-12-13_volume': 'Z',
         '2024-12-14': 'P',
         '2024-12-14_volume': 'AA',
         '2024-12-15': 'Q',
         '2024-12-15_volume': 'AB',
         '2024-12-16': 'R',
         '2024-12-16_volume': 'AC',
         '2024-12-17': 'S',
         '2024-12-17_volume': 'AD'}
        """

        if self.__col_mapping is None:
            self.__col_mapping = {}

            for i in range(7):
                date = self.this_week_start.date() + datetime.timedelta(i)

                self.__col_mapping[f'{date}'] = f'{chr(77 + i)}'

                if i < 3:
                    self.__col_mapping[f'{date}_volume'] = f'{chr(88 + i)}'
                else:
                    self.__col_mapping[f'{date}_volume'] = f'A{chr(65 + i - 3)}'

        return self.__col_mapping

    @property
    def this_week_start(self) -> datetime:
        """
        計算最近一週的起始日期
        """

        return self.specify_day - datetime.timedelta(6)

    @property
    def this_week_end(self):
        return self.specify_day

    @property
    def last_week_start(self):
        """
        計算上週的起始日期
        """

        return self.this_week_start - datetime.timedelta(7)

    @property
    def last_week_end(self):
        """
        計算上週的結束日期
        """

        return self.this_week_end - datetime.timedelta(7)

    @property
    def this_week_date(self) -> List[datetime]:
        """
        生成最近一週的日期區間
        """

        return [self.this_week_start + datetime.timedelta(i) for i in range(7)]

    @property
    def last_week_date(self):
        """
        生成上週的日期區間
        """

        return [self.this_week_start - datetime.timedelta(i) for i in range(8)]

    @property
    def show_monitor(self) -> bool:
        """
        判斷該監控品項是否顯示在報表上，判斷條件為:
            1. 月份是否在監控月份內
            2. 是否為強制(always_display field)顯示
        """

        return self.monitor.months.filter(name=f'{self.specify_day.month}月').exists() or self.monitor.always_display

    @property
    def monitor_has_desc(self) -> bool:
        """
        用於判斷監控品項是否有資料來源說明
        """

        return self.monitor.product.name in self.excel_handler.dict_crop_desc

    @property
    def watchlist(self) -> Watchlist:
        """
        根據指定生成的報表日期，來獲取對應的監控清單
        """

        if self.__watchlist is None:
            self.__watchlist = Watchlist.objects.filter(
                start_date__year=self.specify_day.year,
                start_date__month__lte=self.specify_day.month,
                end_date__month__gte=self.specify_day.month
            ).first()

        return self.__watchlist

    @property
    def monitor_profile_qs(self):
        if self.__monitor_profile_qs is None:
            self.__monitor_profile_qs = MonitorProfile.objects.filter(watchlist=self.watchlist, row__isnull=False)

        return self.__monitor_profile_qs

    @property
    def monitor(self):
        return self.__monitor
    
    @monitor.setter
    def monitor(self, monitor: MonitorProfile):
        self.__monitor = monitor
        
    @property
    def two_weeks_handler(self):
        return self.__two_weeks_handler

    @two_weeks_handler.setter
    def two_weeks_handler(self, handler: DailyTranHandler):
        self.__two_weeks_handler = handler

    @property
    def visible_rows(self):
        return self.__visible_rows

    @visible_rows.setter
    def visible_rows(self, row: int):
        self.__visible_rows.append(row)

    @staticmethod
    def get_simple_avg_price(daily_trans: List[DailyTran]) -> float:
        """
        計算指定日期區間的簡單平均價格
        """

        return (
            sum(d.avg_price for d in daily_trans) / len(daily_trans)
            if daily_trans
            else 0.0
        )

    def extend_query_str(self):
        if self.query.dict_query:
            self.query.dict_query.clear()

        self.query.add_where()

        # 因應措施是梨
        if self.monitor.product.id == 50182:
            # 5, 6 月只抓豐水梨 50186
            if self.specify_day.month in [5, 6]:
                self.query.add_where_by_product_id(50186, has_and=False)
            # 7, 8 月只抓新興梨 50185
            elif self.specify_day.month in [7, 8]:
                self.query.add_where_by_product_id(50185, has_and=False)
            else:
                self.query.add_where_by_product_in(self.monitor.product_list(), has_and=False)
        elif not callable(self.monitor.sources):
            self.query.add_where_by_product_id(self.monitor.product.id, has_and=False)
        else:
            self.query.add_where_by_product_in(self.monitor.product_list(), has_and=False)

        self.query.add_where_by_source_in(self.monitor)
        self.query.save()

    def set_this_week_data(self):
        """
        設定最近一週的平均價格和交易量資料
        """

        self.result[self.monitor.product.name] = {}
        query = self.query.add_where_by_date_between(self.last_week_start, self.this_week_end).build()
        self.two_weeks_handler =  DailyTranHandler(query, self.query.dict_query)

        # 迭代最近兩週的日交易資料，並設定最近一週的平均價格和交易量
        for named_tuple in self.two_weeks_handler.df_with_group_by_date.itertuples():
            # 設定最近一週的平均價格
            if named_tuple.date >= self.this_week_start.date():
                self.result[self.monitor.product.name].update(
                    {f"""{self.col_mapping[f"{named_tuple.date}"]}{self.monitor.row}""": named_tuple.avg_price}
                )

            # 設定最近一週的交易量
            if  self.two_weeks_handler.has_volume and named_tuple.date >= self.this_week_start.date():
                self.result[self.monitor.product.name].update(
                    {f"""{self.col_mapping[f"{named_tuple.date}_volume"]}{self.monitor.row}""": named_tuple.sum_volume}
                )

    def set_avg_price_values(self):
        """
        設定前一週、最近一週的平均價格和與前一週比較的百分比
        """

        last_week_avg_price = self.two_weeks_handler.get_avg_price(
            self.last_week_start.date(), self.last_week_end.date()
        )
        this_week_avg_price = self.two_weeks_handler.get_avg_price(
            self.this_week_start.date(), self.this_week_end.date()
        )

        # '與前一週比較(%)' 欄位(L)
        if last_week_avg_price > 0:
            self.result[self.monitor.product.name].update(
                {f'L{self.monitor.row}': (this_week_avg_price - last_week_avg_price) / last_week_avg_price * 100}
            )

        # 當週(H) 與 前一週(W) '平均價格' 欄位
        self.result[self.monitor.product.name].update(
            {f'H{self.monitor.row}': this_week_avg_price, f'W{self.monitor.row}': last_week_avg_price}
        )

    def set_volume_values(self):
        """
        設定前一週、最近一週的平均交易量和與前一週比較的百分比
        """

        if self.two_weeks_handler.has_volume:
            last_week_avg_volume = self.two_weeks_handler.get_avg_volume(
                self.last_week_start.date(), self.last_week_end.date()
            )
            this_week_avg_volume = self.two_weeks_handler.get_avg_volume(
                self.this_week_start.date(), self.this_week_end.date()
            )

            # 當週交易量欄位(T)
            self.result[self.monitor.product.name].update({f'T{self.monitor.row}': this_week_avg_volume})

            if last_week_avg_volume > 0:
                self.result[self.monitor.product.name].update(
                    {
                        # '與前一週比較(%)' 交易量欄位(U)
                        f'U{self.monitor.row}': 
                            (this_week_avg_volume - last_week_avg_volume) / last_week_avg_volume * 100
                    }
                )

    def set_monitor_price(self):
        """
        設定監控價格欄位
        """

        # 若有監控價格，則設定監控價格欄位(F)
        if self.monitor.price:
            self.result[self.monitor.product.name].update({f'F{self.monitor.row}': self.monitor.price})

    def set_same_month_of_last_year_value(self):
        """
        設定與去年同月份的平均價格
        """

        query = self.query.load().add_where_by_last_year_and_month(self.specify_day).build()
        handler = DailyTranHandler(query, self.query.dict_query)
        last_year_avg_price = handler.get_avg_price()

        # N-1 年 M 月平均價格欄位(G)
        if last_year_avg_price > 0:
            self.result[self.monitor.product.name].update({f'G{self.monitor.row}': last_year_avg_price})

    def update_ram_data(self):
        """
        更新羊資料格式與價格，更新後的格式大概如下:
        {
            'M117': 364.0,
            'N117': 358.0,
            'O117': '(11/07)',
            'P117': '(11/07)',
            'Q117': '(11/07)',
            'R117': 367.0,
            'S117': '(11/11)'
        }
        """

        this_week_date = self.this_week_date
        daily_trans = DailyTran.objects.filter_by_date_lte(
            days=this_week_date, products=self.monitor.product_list(), sources=self.monitor.sources()
        )

        for i, d in enumerate(daily_trans):
            # 依索引取得對應位置的日期
            date = this_week_date[i].date()
            key = f"{self.col_mapping[f'{date}']}{self.monitor.row}"

            # 若是日交易第一筆資料或是交易日期與最近一週日期相同，則設定為該交易的平均價，否則則將值設定為日期
            value = d.avg_price if i == 0 or d.date == date else d.date.strftime('(%m/%d)')
            self.result[self.monitor.product.name].update({key: value})

    def update_cattle_data(self):
        """
        更新牛資價格，因牛資料比較特殊，每週只會有一個價格，e.g. 11/4 -> 128, 11/11 -> 130
        每週的第一個價格，將代表整週都是此價格
        """

        this_week_date = self.this_week_date
        last_week_date = self.last_week_date
        this_week_trans = DailyTran.objects.filter_by_date_lte(this_week_date, self.monitor.product_list())
        this_week_avg_price = self.get_simple_avg_price(this_week_trans)
        last_week_trans = DailyTran.objects.filter_by_date_lte(last_week_date, self.monitor.product_list())
        last_week_avg_price = self.get_simple_avg_price(last_week_trans)

        if this_week_avg_price:
            # 設定最近一週每一日平均價格
            for i, d in enumerate(this_week_trans):
                self.result[self.monitor.product.name].update(
                    {f"{self.col_mapping[f'{this_week_date[i].date()}']}{self.monitor.row}": d.avg_price}
                )

            # 設定最近一週平均價格(H)
            self.result[self.monitor.product.name].update({f'H{self.monitor.row}': this_week_avg_price})

        if last_week_avg_price:
            # 設定前一週平均價格(W)
            self.result[self.monitor.product.name].update({f'W{self.monitor.row}': last_week_avg_price})

        if this_week_avg_price and last_week_avg_price:
            # 設定與前一週比較的百分比(L)
            self.result[self.monitor.product.name].update(
                {
                    f'L{self.monitor.row}': (this_week_avg_price - last_week_avg_price) / last_week_avg_price * 100
                }
            )

    def set_visible_rows(self):
        """
        設定該監控品項是否顯示在報表上，其中 row 代表 Excel 的 row index
        """

        if self.monitor.id is None or self.show_monitor:
            self.excel_handler.visible_rows = self.monitor.row
            
    def set_product_desc(self):
        """
        設定該監控品項的資料來源說明，若該品項有來源說明並且不顯示在報表上，則移除該品項的資料來源說明。
        """

        if self.monitor.id and self.monitor_has_desc and not self.show_monitor:
            self.excel_handler.remove_crop_desc(self.monitor.product.name)

    def report(self):
        for monitor in list(self.monitor_profile_qs) + ExtraItem.get_extra_monitors():
            self.monitor = monitor
            self.extend_query_str()
            self.set_this_week_data()
            self.set_avg_price_values()
            self.set_volume_values()
            self.set_monitor_price()
            self.set_same_month_of_last_year_value()

            if '羊' in monitor.product.name:
                self.update_ram_data()
            if '牛' in monitor.product.name:
                self.update_cattle_data()

            self.set_visible_rows()
            self.set_product_desc()
