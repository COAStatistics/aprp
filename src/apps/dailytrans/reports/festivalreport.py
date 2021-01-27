import pandas as pd 
import numpy as np
import psycopg2
import sxtwl #陽曆陰曆轉換套件
from datetime import datetime,timedelta,date
from sqlalchemy import create_engine
from _pydecimal import Decimal, Context, ROUND_HALF_UP
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from pathlib import Path
from django.conf import settings
from apps.configs.models import FestivalItems, FestivalName, AbstractProduct

#Django ORM 模式
# from apps.dailytrans.models import DailyTran
# from django.db.models import Q

#連結資料庫相關設定
myDBname = settings.DATABASES['default']['NAME']
my_uesrname = settings.DATABASES['default']['USER']
my_pwd = settings.DATABASES['default']['PASSWORD']
my_host = settings.DATABASES['default']['HOST']
my_port = settings.DATABASES['default']['PORT']

con = psycopg2.connect(database=myDBname, user=my_uesrname, password=my_pwd, host=my_host, port=my_port) 
engine = create_engine('postgresql://'+my_uesrname+':'+my_pwd+'@'+my_host+':'+str(my_port)+'/'+myDBname, echo=False) 


class FestivalReportFactory(object):
    def __init__(self, rocyear=date.today().year-1911, festival='1', oneday=False, custom_search=False, custom_search_item=list(), special_day=date.today()-timedelta(days=1)):
        self.roc_year = int(rocyear)
        self.year = self.roc_year + 1911
        self.result_data = dict()
        self.result_volume = dict()
        self.roc_date_range = dict()
        self.today = date.today()
        self.yesterday = self.today - timedelta(days=1)
        self.all_product_id = set()
        self.product_dict = dict()
        self.custom_search = custom_search
        self.custom_search_item = custom_search_item
        if not self.custom_search:
            self.festival = festival
            self.seafood_id_list = []
            self.Chinese_New_Year_dict = dict() #春節_西元年
            self.Dragon_Boat_Festival_dict = dict() #端午節_西元年
            self.Mid_Autumn_Festival_dict = dict() #中秋節_西元年
            self.ROC_Chinese_New_Year_dict = dict() #春節_民國年
            self.ROC_Dragon_Boat_Festival_dict = dict() #端午節_民國年
            self.ROC_Mid_Autumn_Festival_dict = dict() #中秋節_民國年
            self.pid = FestivalItems.objects.filter(festivalname__id__contains=self.festival)
            self.festivalname = FestivalName.objects.filter(id=self.festival)
            self.lunarmonth = self.festivalname[0].lunarmonth
            self.lunarday = self.festivalname[0].lunarday
            for i in self.pid:
                self.product_dict[i.name]=[]
                for j in i.product_id.all():
                    self.product_dict[i.name].append(j.id)
                    self.all_product_id.add(j.id)
        else:
            self.special_day = special_day
            self.year = self.special_day[:4]
            self.roc_year = int(self.year) - 1911
            for i in self.custom_search_item:
                item_data = AbstractProduct.objects.filter(id=i)
                item_name = item_data[0].name
                item_code = item_data[0].code
                self.product_dict[item_name+'_'+item_code]=[]
                self.product_dict[item_name+'_'+item_code].append(int(i))
                self.all_product_id.add(int(i))
        self.roc_before5years = self.roc_year -5
        self.all_product_id_list = list(self.all_product_id)
        
        self.oneday = oneday
        if self.oneday or self.custom_search:
            self.special_day = special_day
            self.special_day_year = self.special_day.split('-')[0]
            self.Custom_dict = dict() #自訂日期_西元年
            self.ROC_Custom_dict = dict() #自訂日期_民國年
        self.table = ''
        self.all_date_list = list()
        self.unit='(元/公斤)'
        

    def festival_date(self, festival=1):
        # 日曆庫實例化
        lunar = sxtwl.Lunar()
        for y in range(self.roc_before5years,self.roc_year+1):
            if not self.custom_search:
                #陰曆轉陽曆
                lunardate='{0}-{1}-{2}'.format(int(y)+1911, self.lunarmonth, self.lunarday) #農曆春節
                lunardate_list = lunardate.split('-')
                solar_day = lunar.getDayByLunar((int)(lunardate_list[0]),(int)(lunardate_list[1]),(int)(lunardate_list[2])) 
                festivalday="{0}-{1}-{2}".format(solar_day.y,solar_day.m,solar_day.d)
            else:
                festivalday_list = self.special_day.split('-')
                festivalday = "{0}-{1}-{2}".format(int(y)+1911,festivalday_list[1],festivalday_list[2])

            festival_day=datetime.strptime(festivalday, "%Y-%m-%d")
            b1w_end = festival_day - timedelta(days=1) #前一週結束日期
            b1w_start = b1w_end - timedelta(days=6) #前一週開始日期
            b2w_end = b1w_start - timedelta(days=1) #前兩週結束日期
            b2w_start = b2w_end - timedelta(days=6) #前兩週開始日期
            b3w_end = b2w_start - timedelta(days=1) #前三週結束日期
            b3w_start = b3w_end - timedelta(days=6) #前三週開始日期
            b4w_end = b3w_start - timedelta(days=1) #前四週結束日期
            b4w_start = b4w_end - timedelta(days=6) #前四週開始日期
            a1w_start = festival_day #節後一週開始日期暨節日當天
            a1w_end = festival_day + timedelta(days=6) #節後一週結束日期
            a2w_start = a1w_end + timedelta(days=1) #節後二週開始日期天
            a2w_end = a2w_start + timedelta(days=6) #節後二週結束日期
            a3w_start = a2w_end + timedelta(days=1) #節後三週開始日期
            a3w_end = a3w_start + timedelta(days=6) #節後三週結束日期
            a4w_start = a3w_end + timedelta(days=1) #節後四週開始日期
            a4w_end = a4w_start + timedelta(days=6) #節後四週結束日期
            begin_date = b4w_start
            last_date = a4w_end
            self.all_date_list.append([begin_date.date(),last_date.date()])

            # if y == self.roc_year:    #今年度新增最近1日的日期區間
            #     if self.festival ==1:
            #         self.Chinese_New_Year_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date()),'{0}~{1}'.format(self.yesterday,self.yesterday)]
            #     elif self.festival ==2:
            #         self.Dragon_Boat_Festival_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date()),'{0}~{1}'.format(self.yesterday,self.yesterday)]
            #     elif self.festival ==3:
            #         self. Mid_Autumn_Festival_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date()),'{0}~{1}'.format(self.yesterday,self.yesterday)]
            # else:
            if not self.custom_search:
                if self.festival =='1':
                    self.Chinese_New_Year_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date())]
                elif self.festival =='2':
                    self.Dragon_Boat_Festival_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date())]
                elif self.festival =='3':
                    self.Mid_Autumn_Festival_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date())]
            else:
                self.Custom_dict[str(y)]=['{0}~{1}'.format(b4w_start.date(),b4w_end.date()),'{0}~{1}'.format(b3w_start.date(),b3w_end.date()),'{0}~{1}'.format(b2w_start.date(),b2w_end.date()),'{0}~{1}'.format(b1w_start.date(),b1w_end.date()),'{0}~{1}'.format(a1w_start.date(),a1w_end.date()),'{0}~{1}'.format(a2w_start.date(),a2w_end.date()),'{0}~{1}'.format(a3w_start.date(),a3w_end.date()),'{0}~{1}'.format(a4w_start.date(),a4w_end.date())]
        
        if not self.custom_search:
            #新增最近一日的日期區間
            # self.all_date_list.append([self.yesterday,self.yesterday])
            #西元年日期區間轉換為民國年日期區間
            if self.festival == '1':
                self.ROC_Chinese_New_Year_dict = self.year2rocyear(self.Chinese_New_Year_dict)
            elif self.festival == '2':
                self.ROC_Dragon_Boat_Festival_dict = self.year2rocyear(self.Dragon_Boat_Festival_dict)
            elif self.festival == '3':
                self.ROC_Mid_Autumn_Festival_dict = self.year2rocyear(self.Mid_Autumn_Festival_dict)
        else:
            self.ROC_Custom_dict = self.year2rocyear(self.Custom_dict)

    #西元年日期區間轉為民國年日期區間
    def year2rocyear(self, date_range):
        if date_range:
            j=len(date_range.values())-1
            for k,v in date_range.items():
                temp_list=[]
                for i in range(len(v)):
                    d=v[i].replace('-','/').replace('{}/'.format(int(self.year)-j),'')
                    temp_list.append(d)
                j=j-1
                self.roc_date_range[k]=temp_list
        else:
            self.roc_date_range = {}
        return self.roc_date_range

    def get_table(self):

        if self.oneday:
            #Pandas dataframe 模式
            #Preventing SQL Injection Attacks
            table = pd.read_sql_query("select product_id, source_id, avg_price, avg_weight, volume, date from dailytrans_dailytran INNER JOIN unnest(%(all_product_id_list)s) as pid ON pid=dailytrans_dailytran.product_id where date = %(day)s ", params={'all_product_id_list':self.all_product_id_list,'day':self.special_day},con=engine)
            # table = pd.read_sql_query(f"select product_id, source_id, avg_price, avg_weight, volume, date from dailytrans_dailytran INNER JOIN unnest(ARRAY{self.all_product_id_list}) as pid ON pid=dailytrans_dailytran.product_id where date = '{self.special_day}' ",con=engine)

            #Django ORM 模式
            # table = DailyTran.objects.filter(product__in=self.all_product_id_list).filter(date = self.special_day)
        else:
            #Pandas dataframe 模式
            #Preventing SQL Injection Attacks
            table = pd.read_sql_query("select product_id, source_id, avg_price, avg_weight, volume, date from dailytrans_dailytran INNER JOIN unnest(%(all_product_id_list)s) as pid ON pid=dailytrans_dailytran.product_id where ((date between %(all_date_list00)s and %(all_date_list01)s) or (date between %(all_date_list10)s and %(all_date_list11)s) or (date between %(all_date_list20)s and %(all_date_list21)s) or (date between %(all_date_list30)s and %(all_date_list31)s) or (date between %(all_date_list40)s and %(all_date_list41)s) or (date between %(all_date_list50)s and %(all_date_list51)s))", params={'all_product_id_list':self.all_product_id_list,'all_date_list00':self.all_date_list[0][0],'all_date_list01':self.all_date_list[0][1],'all_date_list10':self.all_date_list[1][0],'all_date_list11':self.all_date_list[1][1],'all_date_list20':self.all_date_list[2][0],'all_date_list21':self.all_date_list[2][1],'all_date_list30':self.all_date_list[3][0],'all_date_list31':self.all_date_list[3][1],'all_date_list40':self.all_date_list[4][0],'all_date_list41':self.all_date_list[4][1],'all_date_list50':self.all_date_list[5][0],'all_date_list51':self.all_date_list[5][1]},con=engine)
            # table = pd.read_sql_query(f"select product_id, source_id, avg_price, avg_weight, volume, date from dailytrans_dailytran INNER JOIN unnest(ARRAY{self.all_product_id_list}) as pid ON pid=dailytrans_dailytran.product_id where ((date between '{self.all_date_list[0][0]}' and '{self.all_date_list[0][1]}') or (date between '{self.all_date_list[1][0]}' and '{self.all_date_list[1][1]}') or (date between '{self.all_date_list[2][0]}' and '{self.all_date_list[2][1]}') or (date between '{self.all_date_list[3][0]}' and '{self.all_date_list[3][1]}') or (date between '{self.all_date_list[4][0]}' and '{self.all_date_list[4][1]}') or (date between '{self.all_date_list[5][0]}' and '{self.all_date_list[5][1]}'))",con=engine)

            #Django ORM 模式
            # table = DailyTran.objects.filter(product__in=self.all_product_id_list).filter(Q((date >= self.all_date_list[0][0]) & (date <= self.all_date_list[0][1])) | Q((date >= self.all_date_list[1][0]) & (date <= self.all_date_list[1][1])) | Q((date >= self.all_date_list[2][0]) & (date <= self.all_date_list[2][1])) | Q((date >= self.all_date_list[3][0]) & (date <= self.all_date_list[3][1])) | Q((date >= self.all_date_list[4][0]) & (date <= self.all_date_list[4][1])) | Q((date >= self.all_date_list[5][0]) & (date <= self.all_date_list[5][1])) )
        
        #Pandas dataframe 模式需要此兩行轉換類型
        table['source_id'] = table['source_id'].fillna(0).astype(int)
        table['date'] = table['date'].astype(str)
        return table


    def result(self, product_id, source_id, festival='1'):
        self.result_data[str(product_id)]={}
        self.result_volume[str(product_id)]={}
        if self.oneday:
            self.result_data[str(product_id)][str(self.special_day_year)]=[]

            if source_id:
                #Pandas dataframe 模式
                df_product_id = self.table.query(f'product_id in @product_id and (source_id in @source_id)')

                #Django ORM 模式
                # df_product_id = self.table.filter(product__in=product_id).filter(source__in=source_id)
            else:
                #Pandas dataframe 模式
                df_product_id = self.table.query(f'product_id in @product_id')

                #Django ORM 模式
                # df_product_id = self.table.filter(product__in=product_id)

            #Pandas dataframe 模式
            if df_product_id['avg_price'].any():
                has_volume = df_product_id['volume'].notna().sum() / df_product_id['avg_price'].count() > 0.8
                has_weight = df_product_id['avg_weight'].notna().sum() / df_product_id['avg_price'].count() > 0.8
            else:
                has_volume = False
                has_weight = False

            if has_volume and has_weight:
                avgprice=(df_product_id['avg_price']*df_product_id['avg_weight']*df_product_id['volume']).sum()/(df_product_id['avg_weight']*df_product_id['volume']).sum()
                if 80001 <= int(product_id[0]) < 80005: #羊的交易量不是重量x交易量
                    avgvolume = (df_product_id['volume']).sum()
                elif 70001 <= int(product_id[0]) < 70012 : #毛豬交易量為頭數
                    avgvolume = (df_product_id['volume']).sum()
                else:
                    avgvolume = (df_product_id['avg_weight']*df_product_id['volume']).sum()/(df_product_id['volume']).sum()
            elif has_volume:
                avgprice=(df_product_id['avg_price']*df_product_id['volume']).sum()/df_product_id['volume'].sum()
                avgvolume = (df_product_id['volume']).sum()
            else:
                avgprice=df_product_id['avg_price'].mean()
                avgvolume = np.nan
            
            #Django ORM 模式
            # total_price = list()
            # total_volume = list()
            # if df_product_id.count():
            #     if df_product_id.filter(volume__isnull=False).count() / df_product_id.count() > 0.8:
            #         for q in df_product_id:
            #             total_price.append(q.avg_price * q.volume)
            #             total_volume.append(q.volume)
            #             avgprice = sum(total_price) / sum(total_volume) if len(total_volume) else 0
            #     else:
            #         for q in df_product_id:
            #             total_price.append(q.avg_price)
            #             avgprice = sum(total_price) / len(total_price) if len(total_price) else 0
            # else:
            #     avgprice = 'nan'

            self.result_data[str(product_id)][str(self.special_day_year)].append(float(Context(prec=28, rounding=ROUND_HALF_UP).create_decimal(avgprice)))
            self.result_data[str(product_id)][str(self.special_day_year)].append(float(Context(prec=28, rounding=ROUND_HALF_UP).create_decimal(avgvolume)))

        else:
            for y in range(self.roc_before5years,self.roc_year+1):
                self.result_data[str(product_id)][str(y)]=[]
                self.result_volume[str(product_id)][str(y)]=[]

            for y in range(self.roc_before5years,self.roc_year+1):
                if not self.custom_search:
                    if festival =='1':
                        date_range = self.Chinese_New_Year_dict[str(y)]
                    if festival =='2':
                        date_range = self.Dragon_Boat_Festival_dict[str(y)]
                    if festival =='3':
                        date_range = self.Mid_Autumn_Festival_dict[str(y)]
                else:
                    date_range = self.Custom_dict[str(y)]

                for d in date_range:
                    start_date = datetime.strptime("{}".format(d.split('~')[0]), "%Y-%m-%d").date()
                    end_date = datetime.strptime("{}".format(d.split('~')[1]), "%Y-%m-%d").date()

                    if source_id:
                        df_product_id = self.table.query(f'product_id in @product_id and date >= "{start_date}" and date <= "{end_date}" and (source_id in @source_id)')
                        # df_product_id = self.table.filter(product__in=product_id).filter(source__in=source_id).filter(date >=start_date).filter(date <= end_date)
                        
                    else:
                        df_product_id = self.table.query(f'product_id in @product_id and date >= "{start_date}" and date <= "{end_date}"')
                        # df_product_id = self.table.filter(product__in=product_id).filter(date >=start_date).filter(date <= end_date)

                    if df_product_id.shape[0]:
                        has_volume = df_product_id['volume'].count() / df_product_id.shape[0] > 0.8
                        has_weight = df_product_id['avg_weight'].count() / df_product_id.shape[0] > 0.8
                    else:
                        has_volume = False
                        has_weight = False
                    if has_volume and has_weight:
                        avgprice=(df_product_id['avg_price']*df_product_id['avg_weight']*df_product_id['volume']).sum()/(df_product_id['avg_weight']*df_product_id['volume']).sum()
                        if 80001 <= int(product_id[0]) < 80005: #羊的交易量不是重量x交易量
                            avgvolume = (df_product_id['volume']).sum()
                        elif 70001 <= int(product_id[0]) < 70012 : #毛豬交易量為頭數
                            avgvolume = (df_product_id['volume']).sum()
                        else:
                            avgvolume=(df_product_id['avg_weight']*df_product_id['volume']).sum()/(df_product_id['volume']).sum()
                    elif has_volume:
                        avgprice=(df_product_id['avg_price']*df_product_id['volume']).sum()/df_product_id['volume'].sum()
                        avgvolume=df_product_id['volume'].sum()
                    else:
                        avgprice=df_product_id['avg_price'].mean()
                        avgvolume=np.nan



                    # total_price = list()
                    # total_volume = list()
                    # if df_product_id.count():
                    #     if df_product_id.filter(volume__isnull=False).count() / df_product_id.count() > 0.8:
                    #         for q in df_product_id:
                    #             total_price.append(q.avg_price * q.volume)
                    #             total_volume.append(q.volume)
                    #             avgprice = sum(total_price) / sum(total_volume) if len(total_volume) else 0
                    #     else:
                    #         for q in df_product_id:
                    #             total_price.append(q.avg_price)
                    #             avgprice = sum(total_price) / len(total_price) if len(total_price) else 0
                    # else:
                    #     avgprice = 'nan'
                    self.result_data[str(product_id)][str(y)].append(float(Context(prec=28, rounding=ROUND_HALF_UP).create_decimal(avgprice)))
                    self.result_volume[str(product_id)][str(y)].append(float(Context(prec=28, rounding=ROUND_HALF_UP).create_decimal(avgvolume)))

        return self.result_data,self.result_volume

    def trimmean(self, arr):
        arr_min = arr.min()
        arr_max = arr.max()
        arr_sum = arr.sum()
        arr_trimmean = (arr_sum - arr_max - arr_min)/(len(arr)-2)
        return arr_trimmean

    def data2pandas2save(self, product_dict, result_data, festival='1', volume=0):
        year_df = pd.Series(result_data.values())
        index_df = pd.Series(list(result_data.keys()))
        df1 = pd.DataFrame(index_df,index=[i for i in range(0,len(self.product_dict.keys()))])
        df2 = pd.DataFrame(list(year_df[0]),index=[i for i in range(0,len(self.product_dict.keys()))])
        df3 = pd.merge(df1,df2, left_index =True, right_index =True, how='outer')
        df3.columns.values[0]='農產品'
        for index, row in df3.iterrows():
            product_name = [k for k in list(self.product_dict.keys())][index]
            df3.loc[index,'農產品'] = product_name

        Row_list =[]
        for index, rows in df3.iterrows(): 
            data_list = []
            for y in range(self.roc_before5years,self.roc_year+1):
                data_list.append(rows[str(y)])
            l = [y for x in data_list for y in x]
            Row_list.append(l)

        columns=[]
        for y in range(self.roc_before5years,self.roc_year+1):
            #最近一日
            # if y == self.roc_year:
            #     for i in range(0,9):
            #         temp_list=[]
            #         if i ==0:
            #             temp='節前四週'
            #         elif i ==1:
            #             temp='節前三週'
            #         elif i ==2:
            #             temp='節前二週'
            #         elif i ==3:
            #             temp='節前一週'
            #         elif i ==4:
            #             temp='節後一週'
            #         elif i ==5:
            #             temp='節後二週'
            #         elif i ==6:
            #             temp='節後三週'
            #         elif i ==7:
            #             temp='節後四週'
            #         elif i ==8:
            #             temp='最近一日'
            #         temp_list.append(str(y))
            #         temp_list.append(temp)
            #         columns.append(tuple(temp_list))
            # else:
            for i in range(0,8):
                temp_list=[]
                if i ==0:
                    temp='節前四週'
                elif i ==1:
                    temp='節前三週'
                elif i ==2:
                    temp='節前二週'
                elif i ==3:
                    temp='節前一週'
                elif i ==4:
                    temp='節後一週'
                elif i ==5:
                    temp='節後二週'
                elif i ==6:
                    temp='節後三週'
                elif i ==7:
                    temp='節後四週'
                temp_list.append(str(y))
                temp_list.append(temp)
                columns.append(tuple(temp_list))

        df4 = pd.DataFrame(Row_list,columns=pd.MultiIndex.from_tuples(columns))

        for index, row in df4.iterrows():
            product_name = [k for k in list(self.product_dict.keys())][index]
            df4.loc[index,('農產品','產品名稱')] = product_name

        if volume:
            #不顯示交易量的單位
            self.unit=''


        for index, row in df4.iterrows():
            df4.loc[index,('節前四週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節前四週')]/df4.loc[index,(str(self.roc_year-1), '節前四週')]*100-100
            df4.loc[index,('節前四週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節前四週')]-df4.loc[index,(str(self.roc_year-1), '節前四週')]
            df4.loc[index,('前五年簡單平均','節前四週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節前四週'),(str(self.roc_year-2), '節前四週'),(str(self.roc_year-3), '節前四週'),(str(self.roc_year-4), '節前四週'),(str(self.roc_year-5), '節前四週')]])

            df4.loc[index,('節前三週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節前三週')]/df4.loc[index,(str(self.roc_year-1), '節前三週')]*100-100
            df4.loc[index,('節前三週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節前三週')]-df4.loc[index,(str(self.roc_year-1), '節前三週')]
            df4.loc[index,('前五年簡單平均','節前三週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節前三週'),(str(self.roc_year-2), '節前三週'),(str(self.roc_year-3), '節前三週'),(str(self.roc_year-4), '節前三週'),(str(self.roc_year-5), '節前三週')]])

            df4.loc[index,('節前二週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節前二週')]/df4.loc[index,(str(self.roc_year-1), '節前二週')]*100-100
            df4.loc[index,('節前二週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節前二週')]-df4.loc[index,(str(self.roc_year-1), '節前二週')]
            df4.loc[index,('前五年簡單平均','節前二週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節前二週'),(str(self.roc_year-2), '節前二週'),(str(self.roc_year-3), '節前二週'),(str(self.roc_year-4), '節前二週'),(str(self.roc_year-5), '節前二週')]])

            df4.loc[index,('節前一週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節前一週')]/df4.loc[index,(str(self.roc_year-1), '節前一週')]*100-100
            df4.loc[index,('節前一週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節前一週')]-df4.loc[index,(str(self.roc_year-1), '節前一週')]
            df4.loc[index,('前五年簡單平均','節前一週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節前一週'),(str(self.roc_year-2), '節前一週'),(str(self.roc_year-3), '節前一週'),(str(self.roc_year-4), '節前一週'),(str(self.roc_year-5), '節前一週')]])

            df4.loc[index,('節後一週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節後一週')]/df4.loc[index,(str(self.roc_year-1), '節後一週')]*100-100
            df4.loc[index,('節後一週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節後一週')]-df4.loc[index,(str(self.roc_year-1), '節後一週')]
            df4.loc[index,('前五年簡單平均','節後一週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節後一週'),(str(self.roc_year-2), '節後一週'),(str(self.roc_year-3), '節後一週'),(str(self.roc_year-4), '節後一週'),(str(self.roc_year-5), '節後一週')]])

            df4.loc[index,('節後二週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節後二週')]/df4.loc[index,(str(self.roc_year-1), '節後二週')]*100-100
            df4.loc[index,('節後二週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節後二週')]-df4.loc[index,(str(self.roc_year-1), '節後二週')]
            df4.loc[index,('前五年簡單平均','節後二週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節後二週'),(str(self.roc_year-2), '節後二週'),(str(self.roc_year-3), '節後二週'),(str(self.roc_year-4), '節後二週'),(str(self.roc_year-5), '節後二週')]])

            df4.loc[index,('節後三週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節後三週')]/df4.loc[index,(str(self.roc_year-1), '節後三週')]*100-100
            df4.loc[index,('節後三週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節後三週')]-df4.loc[index,(str(self.roc_year-1), '節後三週')]
            df4.loc[index,('前五年簡單平均','節後三週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節後三週'),(str(self.roc_year-2), '節後三週'),(str(self.roc_year-3), '節後三週'),(str(self.roc_year-4), '節後三週'),(str(self.roc_year-5), '節後三週')]])

            df4.loc[index,('節後四週與去年同期比較','漲跌率(%)')] = df4.loc[index,(str(self.roc_year), '節後四週')]/df4.loc[index,(str(self.roc_year-1), '節後四週')]*100-100
            df4.loc[index,('節後四週與去年同期比較','差幅{}'.format(self.unit))] = df4.loc[index,(str(self.roc_year), '節後四週')]-df4.loc[index,(str(self.roc_year-1), '節後四週')]
            df4.loc[index,('前五年簡單平均','節後四週平均{}'.format(self.unit))] = self.trimmean(df4.loc[index,[(str(self.roc_year-1), '節後四週'),(str(self.roc_year-2), '節後四週'),(str(self.roc_year-3), '節後四週'),(str(self.roc_year-4), '節後四週'),(str(self.roc_year-5), '節後四週')]])

        df4.index = df4.index + 1
        columns_list=[]
        columns_list.append(('農產品','產品名稱'))
        # columns_list.append(('{}'.format(self.roc_year),'最近一日'))
        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節前四週'))
        columns_list.append(('節前四週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節前四週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節前四週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節前三週'))
        columns_list.append(('節前三週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節前三週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節前三週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節前二週'))
        columns_list.append(('節前二週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節前二週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節前二週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節前一週'))
        columns_list.append(('節前一週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節前一週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節前一週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節後一週'))
        columns_list.append(('節後一週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節後一週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節後一週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節後二週'))
        columns_list.append(('節後二週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節後二週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節後二週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節後三週'))
        columns_list.append(('節後三週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節後三週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節後三週平均{}'.format(self.unit)))

        for y in range(self.roc_year,self.roc_year-5-1,-1):
            columns_list.append(('{}'.format(y), '節後四週'))
        columns_list.append(('節後四週與去年同期比較', '漲跌率(%)'))
        columns_list.append(('節後四週與去年同期比較', '差幅{}'.format(self.unit)))
        columns_list.append(('前五年簡單平均', '節後四週平均{}'.format(self.unit)))

        df4=df4[columns_list]

        df5 = df4.copy()
        if not self.custom_search:
            if festival =='1':
                self.roc_date_range = self.ROC_Chinese_New_Year_dict
            if festival =='2':
                self.roc_date_range = self.ROC_Dragon_Boat_Festival_dict
            if festival =='3':
                self.roc_date_range = self.ROC_Mid_Autumn_Festival_dict
        else:
            self.roc_date_range = self.ROC_Custom_dict

        #組合新欄位名稱
        yes_date=self.roc_date_range[str(self.roc_year)][-1].split('~')[0]
        if len(yes_date) > 5:
            yes_date = yes_date.split('/')[1]+'/'+yes_date.split('/')[2]
        columns_name = []
        columns_name.append('農產品')
        # columns_name.append('最近一日\n{}/{}\n(元/公斤)'.format(self.today.year-1911,yes_date))
        if not self.custom_search:
            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][0]
                columns_name.append('{0}年節前四週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前四週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前四週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節前四週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][1]
                columns_name.append('{0}年節前三週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前三週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前三週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節前三週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][2]
                columns_name.append('{0}年節前二週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前二週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前二週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節前二週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][3]
                columns_name.append('{0}年節前一週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前一週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前一週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節前一週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][4]
                columns_name.append('{0}年節後一週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後一週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後一週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節後一週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][5]
                columns_name.append('{0}年節後二週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後二週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後二週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節後二週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][6]
                columns_name.append('{0}年節後三週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後三週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後三週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節後三週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][7]
                columns_name.append('{0}年節後四週\n{1}\n{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後四週\n較{}年同期\n漲跌率\n(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後四週\n較{}年同期\n差幅\n{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均\n({}-{}年)\n節後四週\n{}'.format(self.roc_year-5,self.roc_year-1,self.unit))
        else:
            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][0]
                columns_name.append('{0}年節前四週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前四週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前四週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節前四週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][1]
                columns_name.append('{0}年節前三週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前三週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前三週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節前三週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][2]
                columns_name.append('{0}年節前二週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前二週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前二週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節前二週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][3]
                columns_name.append('{0}年節前一週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節前一週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節前一週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節前一週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][4]
                columns_name.append('{0}年節後一週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後一週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後一週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節後一週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][5]
                columns_name.append('{0}年節後二週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後二週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後二週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節後二週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][6]
                columns_name.append('{0}年節後三週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後三週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後三週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節後三週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

            for y in range(self.roc_year,self.roc_year-6,-1):
                date=self.roc_date_range[str(y)][7]
                columns_name.append('{0}年節後四週<br>{1}<br>{2}'.format(y,date,self.unit))
            columns_name.append('{}年節後四週<br>較{}年同期<br>漲跌率<br>(%)'.format(self.roc_year,self.roc_year-1))
            columns_name.append('{}年節後四週<br>較{}年同期<br>差幅<br>{}'.format(self.roc_year,self.roc_year-1,self.unit))
            columns_name.append('近5年簡單平均<br>({}-{}年)<br>節後四週<br>{}'.format(self.roc_year-5,self.roc_year-1,self.unit))

        df5.columns = columns_name
        if not self.custom_search:
            festival_name = FestivalName.objects.filter(id = self.festival)
            festival_title = festival_name[0].name
            if volume:
                file_name = '{}_{}節前交易表.xlsx'.format(self.roc_year,festival_title)
            else:
                file_name = '{}_{}節前價格表.xlsx'.format(self.roc_year,festival_title)

            writer = pd.ExcelWriter(file_name)
            wb = openpyxl.Workbook()
            df6 = df5.copy()
            df6.fillna('-').to_excel(writer)
            if volume:
                ws = wb.create_sheet(index=0, title="交易量表")
            else:
                ws = wb.create_sheet(index=0, title="價格表")

            #pandas to openpyxl
            for r in dataframe_to_rows(df6, index=False, header=False):
                ws.append(r)

            #設定字形大小及格式
            title_font = Font(size=22)
            content_font = Font(size=16)
            alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

            #新增框線
            border = Border(top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
            for _row in ws.iter_rows():
                for _cell in _row:
                    _cell.border = border
                    _cell.font = content_font
                    _cell.number_format = '#,##0.0' #小數一位

            #在第一行前插入一行
            ws.insert_rows(1)
            # 合併儲存格
            ws.merge_cells('A1:BU1')
            #儲存格內容及格式
            if volume:
                ws['A1']='{}節前農產品交易量變動情形表'.format(festival_title)
            else:
                ws['A1']='{}節前農產品價格變動情形表'.format(festival_title)
            ws['A1'].alignment=Alignment(horizontal='center', vertical='center')
            ws['A1'].font = title_font

            #插入兩行製作各欄位標題
            ws.insert_rows(2)
            ws.insert_rows(2)
            for i in range(1,len(df6.columns)+1):
                if i in [8,9,17,18,26,27,35,36,44,45,53,54,62,63,71,72]:
                    ws.cell(row=2, column=i).value=df6.columns[i-1].split('\n')[0]
                    ws.cell(row=3, column=i).value=df6.columns[i-1].split('\n')[1]+'\n'+df6.columns[i-1].split('\n')[2]
                else:
                    ws.cell(row=2, column=i).value=df6.columns[i-1]
                ws.cell(row=2, column=i).font=content_font
                ws.cell(row=3, column=i).font=content_font
                

            # 合併儲存格
            #垂直合併
            for i in range(65,91):#A~Z
                if i in (72,73,81,82,90):#H,I,Q,R,Z
                    continue
                ws.merge_cells('{0}2:{0}3'.format(chr(i))) 
            for i in range(65,91):
                if i in (65,73,74,82,83):#A,I,J,R,S
                    continue
                ws.merge_cells('A{0}2:A{0}3'.format(chr(i)))
            for i in range(65,91):
                if i in (65,66,74,75,83,84):#A,B,J,K,S,T
                    continue
                ws.merge_cells('B{0}2:B{0}3'.format(chr(i)))
            #水平合併
            ws.merge_cells('H2:I2')
            ws.merge_cells('Q2:R2')
            ws.merge_cells('Z2:AA2')
            ws.merge_cells('AI2:AJ2')
            ws.merge_cells('AR2:AS2')
            ws.merge_cells('BA2:BB2')
            ws.merge_cells('BJ2:BK2')
            ws.merge_cells('BS2:BT2')

            # 設定儲存格格式
            for i in range(1,len(df6.columns)+1):
                ws.cell(row=2, column=i).alignment=alignment    #置中及換行
                ws.cell(row=3, column=i).alignment=alignment
                ws.cell(row=2, column=i).border=border  #框線
                ws.cell(row=3, column=i).border=border

            #欄位寬度
            ws.column_dimensions['A'].width = 33
            for c in range(2,len(df6.columns)+1):
                ws.column_dimensions[get_column_letter(c)].width = 22
            for _col in ['H','I','Q', 'R','Z','AA','AI','AJ','AR','AS','BA','BB','BJ','BK','BS','BT']:
                ws.column_dimensions[_col].width = 18.5

            #設定第2列高度
            ws.row_dimensions[2].height =42

            #隱藏欄位
            # hidden_columns=['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO']
            # for col in hidden_columns:
            #     ws.column_dimensions[col].hidden= True

            #凍結窗格
            ws.freeze_panes = ws['B4']

            # Close the Pandas Excel writer and output the Excel file.
            wb.save(file_name)
            return file_name
        else:
            df6 = df5.round(1)
            return df6

    def __call__(self, output_dir=settings.BASE_DIR):
        #產生節日日期區間
        if not self.custom_search:
            self.festival_date(self.festival)
        else:
            self.festival_date()
        #獲取完整交易表
        self.table = self.get_table()

        if not self.custom_search:
            for i in self.pid:
                product_id_list = []
                source_id_list = []
                for j in i.product_id.all():
                    product_id_list.append(j.id)

                for k in i.source.all():
                    source_id_list.append(k.id)

                self.result_data,self.result_volume = self.result(product_id_list, source_id_list, self.festival)
 
            #只查詢單一日期,返回各品項查詢值
            if self.oneday:
                return self.result_data
                
            #產生八週週期日報,返回日報名稱
            else:
                file_name = self.data2pandas2save(self.product_dict, self.result_data, self.festival)
                file_path = Path(output_dir, file_name)
                file_volume_name = self.data2pandas2save(self.product_dict, self.result_volume, self.festival,volume=1)
                file_volume_path = Path(output_dir, file_volume_name)
            return file_name, file_path, file_volume_name, file_volume_path

        else:
            for i in self.custom_search_item:
                product_id_list = []
                product_id_list.append(i)
                
                #蔬菜(40001~49011),批發蔬菜(10006~10010),來源為[10001, 10002]
                #水果(50001~59019),批發水果(20001~20004),來源為[20001, 20002]
                #花卉(60001~60100),批發花卉(30001~30002),來源為[30001]
                #魚(120001~121048),來源為[80001, 80003, 80006, 80007, 80008, 80009, 80010, 80011, 80013, 80014, 80017, 80018, 80020]
                #羊(80001~80004),來源為[50001] 彰化市場
                #毛豬(70001~70009),來源為[40002, 40003, 40004, 40005, 40006, 40007, 40008, 40009, 40010, 40011, 40012, 40013, 40014, 40015, 40016, 40017, 40018, 40019, 40020, 40021,40022, 40023],台灣地區不含澎湖
                #糧(1~27),雞(90004~90016),鴨(100001~100006),雞(110001~110006),牛(130001~130005),沒有來源
                if 40001 <= int(i) < 50000 or 10006 <= int(i) <= 10010: #蔬菜類,批發蔬菜
                    source_id_list = [10001, 10002]
                elif 50001 <= int(i) < 60000 or 20001 <= int(i) <= 20004: #水果,批發水果
                    source_id_list = [20001, 20002]
                elif 60001 <= int(i) < 70000 or 30001 <= int(i) <= 30002: #花卉,批發花卉
                    source_id_list = [30001]
                elif 120001 <= int(i) < 130000: #漁產品
                    source_id_list = [80001, 80003, 80006, 80007, 80008, 80009, 80010, 80011, 80013, 80014, 80017, 80018, 80020]
                elif 80001 <= int(i) < 80005: #羊,抓彰化市場
                    source_id_list = [50001]
                elif 70001 <= int(i) < 70012: #毛豬,台灣地區不含澎湖
                    source_id_list = [40002, 40003, 40004, 40005, 40006, 40007, 40008, 40009, 40010, 40011, 40012, 40013, 40014, 40015, 40016, 40017, 40018, 40019, 40020, 40021, 40022]
                else:
                    source_id_list = None

                self.result_data,self.result_volume = self.result(product_id = i, source_id = source_id_list)

            product_dataframe = self.data2pandas2save(self.product_dict, self.result_data)
            product_dataframe_volume = self.data2pandas2save(self.product_dict, self.result_volume, volume=1)
            return product_dataframe, product_dataframe_volume

        #各品項逐一查詢價格
        # for k,v in self.product_dict.items():
            # self.result_data = self.get_data(self.festival,v)
